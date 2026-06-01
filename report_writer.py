"""
AACP Lab v3 - Report Writer
Multi-sheet Excel report covering all 5 workflows.
One workbook per run. Formatted with RAG status colouring.
"""

from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

COLOURS = {
    "header_bg":  "1F4E79",
    "header_fg":  "FFFFFF",
    "breach_bg":  "FFCCCC",
    "warning_bg": "FFF2CC",
    "ok_bg":      "E2EFDA",
    "green_bg":   "E2EFDA",
    "red_bg":     "FFCCCC",
    "amber_bg":   "FFF2CC",
    "total_bg":   "F2F2F2",
    "section_bg": "D6E4F0",
}

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, colour="000000", size=11): return Font(bold=bold, color=colour, size=size)
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _centre(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _wrap():   return Alignment(wrap_text=True, vertical="top")

def hdr_row(ws, row, cols, widths):
    for i, (col, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font      = _font(bold=True, colour=COLOURS["header_fg"])
        c.fill      = _fill(COLOURS["header_bg"])
        c.border    = _border()
        c.alignment = _centre()
        ws.column_dimensions[get_column_letter(i)].width = w

def cell(ws, row, col, value, bg=None, bold=False, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.border    = _border()
    c.alignment = _wrap() if wrap else _centre()
    if bg: c.fill = _fill(bg)
    if bold: c.font = _font(bold=True)
    return c


def write_report(results: dict, model: str, period: str, ts: str) -> str:
    if not HAS_OPENPYXL:
        print("  openpyxl not installed — skipping Excel")
        return None

    Path("output").mkdir(exist_ok=True)
    path = f"output/lab_v3_{period}_{ts}.xlsx"
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Summary ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 12

    title = ws1.cell(row=1, column=1, value=f"AACP LAB v3 — {period} — {model}")
    title.font = _font(bold=True, colour=COLOURS["header_fg"], size=13)
    title.fill = _fill(COLOURS["header_bg"])
    ws1.merge_cells("A1:E1")
    ws1.row_dimensions[1].height = 24

    hdr_row(ws1, 2, ["Workflow", "Hops", "Tokens In", "Cost (USD)", "Success"], [26,8,12,14,10])

    wf_labels = {
        "payroll":  "Payroll Q1 FY2026",
        "month_end": "Month-End Close",
        "sales":    "Sales Qualification",
        "cs":       "CS Resolution",
        "jml":      "JML Onboarding",
    }
    total_cost = total_tokens = total_hops = 0
    for row_n, (key, r) in enumerate(results.items(), start=3):
        t = r.get("totals", {})
        s = "Yes" if r.get("success") else "No"
        bg = COLOURS["ok_bg"] if r.get("success") else COLOURS["breach_bg"]
        cell(ws1, row_n, 1, wf_labels.get(key, key), bg)
        cell(ws1, row_n, 2, t.get("hops", 0), bg)
        cell(ws1, row_n, 3, t.get("tokens_in", 0), bg)
        cell(ws1, row_n, 4, f"${t.get('cost_usd',0):.4f}", bg)
        cell(ws1, row_n, 5, s, bg)
        total_hops   += t.get("hops", 0)
        total_tokens += t.get("tokens_in", 0)
        total_cost   += t.get("cost_usd", 0)

    tr = len(results) + 3
    for col, val in enumerate([f"TOTAL ({len(results)} workflows)",
                                total_hops, f"{total_tokens:,}",
                                f"${total_cost:.4f}", ""], start=1):
        cell(ws1, tr, col, val, COLOURS["total_bg"], bold=True)

    # ── Sheet 2: Payroll ───────────────────────────────────────────────────
    if "payroll" in results:
        ws2 = wb.create_sheet("Payroll")
        pr  = results["payroll"]
        report = (pr.get("output") or {}).get("report") or {}
        payroll_data = (pr.get("output") or {}).get("payroll") or {}

        ws2.cell(row=1, column=1, value=f"PAYROLL — {period}").font = _font(bold=True, size=12)
        ws2.cell(row=2, column=1, value=report.get("executive_summary",""))
        ws2.merge_cells("A2:G2")
        ws2.row_dimensions[2].height = 50
        ws2.cell(row=2, column=1).alignment = _wrap()

        hdr_row(ws2, 4, ["ID","Name","Dept","Cost Centre","Gross","PAYE","Net","Status"],
                [8,20,16,14,14,14,14,12])
        employees = (payroll_data.get("employees") or [])
        for row_n, emp in enumerate(employees, start=5):
            status = emp.get("budget_status","OK")
            bg = COLOURS["breach_bg"] if status=="BREACH" else \
                 COLOURS["warning_bg"] if status=="WARNING" else None
            for col, val in enumerate([
                emp.get("id"), emp.get("name"), emp.get("dept"),
                emp.get("cost_centre"), f"£{int(emp.get('gross_pay') or 0):,}",
                f"£{int(emp.get('paye') or 0):,}", f"£{int(emp.get('net_pay') or 0):,}",
                status
            ], start=1):
                cell(ws2, row_n, col, val, bg)

        totals = (payroll_data.get("totals") or {})
        tr2 = len(employees) + 5
        for col, val in enumerate(["","TOTALS","","",
            f"£{int(totals.get('gross') or 0):,}",
            f"£{int(totals.get('paye') or 0):,}",
            f"£{int(totals.get('net') or 0):,}",""], start=1):
            cell(ws2, tr2, col, val, COLOURS["total_bg"], bold=True)

        # Anomalies
        anomalies = report.get("anomalies") or []
        if anomalies:
            ws2.cell(row=tr2+2, column=1, value="ANOMALIES").font = _font(bold=True)
            hdr_row(ws2, tr2+3, ["Severity","Employee","CC","Detail","Action"], [10,20,14,40,36])
            for row_n, an in enumerate(anomalies, start=tr2+4):
                bg = COLOURS["breach_bg"] if an.get("severity")=="BREACH" else COLOURS["warning_bg"]
                for col, val in enumerate([
                    an.get("severity"), an.get("employee"),
                    an.get("cc"), an.get("detail"), an.get("action")
                ], start=1):
                    cell(ws2, row_n, col, val, bg)
                ws2.row_dimensions[row_n].height = 36

    # ── Sheet 3: Sales Pipeline ────────────────────────────────────────────
    if "sales" in results:
        ws3 = wb.create_sheet("Sales Pipeline")
        sr  = results["sales"]
        leads = (sr.get("output") or {}).get("results") or []
        qualified = (sr.get("output") or {}).get("qualified", 0)

        ws3.cell(row=1, column=1,
                 value=f"SALES QUALIFICATION — {period} — {len(leads)} leads, {qualified} qualified"
                 ).font = _font(bold=True, size=12)

        hdr_row(ws3, 3, ["Lead ID","Company","Score","Qualified","Stage","Rep","BANT Budget","BANT Need"],
                [10,24,10,12,20,20,14,12])
        for row_n, lead in enumerate(leads, start=4):
            qual = lead.get("qualified")
            bg   = COLOURS["ok_bg"] if qual else COLOURS["amber_bg"]
            bant = lead.get("bant") or {}
            for col, val in enumerate([
                lead.get("lead_id"), lead.get("company"),
                lead.get("score"), "Yes" if qual else "No",
                lead.get("stage"), lead.get("rep"),
                bant.get("budget"), bant.get("need")
            ], start=1):
                cell(ws3, row_n, col, val, bg)

    # ── Sheet 4: CS Resolution ─────────────────────────────────────────────
    if "cs" in results:
        ws4 = wb.create_sheet("CS Resolution")
        cr  = results["cs"]
        tickets = (cr.get("output") or {}).get("results") or []
        gw_count = (cr.get("output") or {}).get("goodwill_offered", 0)

        ws4.cell(row=1, column=1,
                 value=f"CS RESOLUTION — {period} — {len(tickets)} tickets, {gw_count} goodwill offers"
                 ).font = _font(bold=True, size=12)

        hdr_row(ws4, 3, ["Ticket","Customer","LTV","Category","Strategy","Goodwill","Amount","Sent"],
                [10,12,10,14,36,10,10,8])
        for row_n, t in enumerate(tickets, start=4):
            gw = t.get("goodwill_offer")
            bg = COLOURS["ok_bg"] if t.get("sent") else COLOURS["amber_bg"]
            for col, val in enumerate([
                t.get("ticket_id"), t.get("customer_id"),
                f"£{int(t.get('ltv_gbp') or 0):,}",
                t.get("category"),
                t.get("resolution_strategy"),
                "Yes" if gw else "No",
                f"£{int(t.get('goodwill_amount') or 0):,}" if gw else "-",
                "Yes" if t.get("sent") else "No",
            ], start=1):
                cell(ws4, row_n, col, val, bg)
            ws4.row_dimensions[row_n].height = 36

    # ── Sheet 5: JML Onboarding ────────────────────────────────────────────
    if "jml" in results:
        ws5 = wb.create_sheet("JML Onboarding")
        jr  = results["jml"]
        hires = (jr.get("output") or {}).get("results") or []
        prov  = (jr.get("output") or {}).get("provisioned", 0)

        ws5.cell(row=1, column=1,
                 value=f"JML ONBOARDING — {period} — {len(hires)} new hires, {prov} provisioned"
                 ).font = _font(bold=True, size=12)

        hdr_row(ws5, 3, ["ID","Name","Username","Dept","Role","Email","Licences","Systems","Welcome"],
                [8,20,14,14,20,28,24,24,10])
        for row_n, h in enumerate(hires, start=4):
            bg = COLOURS["ok_bg"] if h.get("welcome_sent") else COLOURS["amber_bg"]
            licences = ", ".join(h.get("licences") or []) if isinstance(h.get("licences"), list) else str(h.get("licences",""))
            systems  = ", ".join(h.get("systems")  or []) if isinstance(h.get("systems"),  list) else str(h.get("systems",""))
            for col, val in enumerate([
                h.get("employee_id"), h.get("name"), h.get("username"),
                h.get("dept"), h.get("role"), h.get("email"),
                licences, systems,
                "Yes" if h.get("welcome_sent") else "No",
            ], start=1):
                cell(ws5, row_n, col, val, bg)

    # ── Sheet 6: Month-End Close ───────────────────────────────────────────
    if "month_end" in results:
        ws6 = wb.create_sheet("Month-End Close")
        mr  = results["month_end"]
        mgmt = (mr.get("output") or {}).get("management_accounts") or {}
        recon = (mr.get("output") or {}).get("reconciliation") or {}
        var   = (mr.get("output") or {}).get("variance_analysis") or {}

        ws6.cell(row=1, column=1,
                 value=f"MONTH-END CLOSE — {period}").font = _font(bold=True, size=12)
        ws6.cell(row=2, column=1,
                 value=mgmt.get("executive_summary",""))
        ws6.merge_cells("A2:F2")
        ws6.row_dimensions[2].height = 50

        ws6.cell(row=4, column=1, value="RECONCILIATION").font = _font(bold=True)
        hdr_row(ws6, 5, ["Status","Matched","Unmatched","Unmatched Value","Reconciled"], [14,12,12,18,12])
        cell(ws6, 6, 1, recon.get("status",""))
        cell(ws6, 6, 2, recon.get("matched_items",0))
        cell(ws6, 6, 3, recon.get("unmatched_items",0))
        cell(ws6, 6, 4, f"£{int(recon.get('unmatched_value_gbp') or 0):,}")
        cell(ws6, 6, 5, "Yes" if recon.get("reconciled") else "No",
             COLOURS["ok_bg"] if recon.get("reconciled") else COLOURS["breach_bg"])

        variances = var.get("variances") or []
        if variances:
            ws6.cell(row=8, column=1, value="VARIANCES").font = _font(bold=True)
            hdr_row(ws6, 9, ["Category","Current","Prior","Variance","Var %","Material","Note"],
                    [20,14,14,14,10,10,30])
            for row_n, v in enumerate(variances, start=10):
                bg = COLOURS["breach_bg"] if v.get("material") else None
                for col, val in enumerate([
                    v.get("category"),
                    f"£{int(v.get('current_gbp') or 0):,}",
                    f"£{int(v.get('prior_gbp') or 0):,}",
                    f"£{int(v.get('variance_gbp') or 0):,}",
                    f"{float(v.get('variance_pct') or 0):.1f}%",
                    "Yes" if v.get("material") else "No",
                    v.get("note",""),
                ], start=1):
                    cell(ws6, row_n, col, val, bg)

    wb.save(path)
    return path
